# -*- coding: utf-8 -*-
import pulp
import openpyxl


def get_data_from_page(req):
    data_dict = req.to_dict()
    b_num = len({k for k, v in data_dict.items() if 'max' in k})
    x_num = len({k for k, v in data_dict.items() if 'x_name_' in k})
    x = {}
    max_min = {}
    for i in range(0, x_num):
        x_name = chr(ord('a')+i)
        x[x_name] = {}
        for j in range(0, b_num):
            b_index = str(j+1)
            if req[x_name+b_index]:
                x[x_name][str(j+1)] = float(req[x_name+b_index])
            else:
                x[x_name][str(j + 1)] = 0.0
            if req[x_name+'_cost']:
                x[x_name]['cost'] = float(req[x_name+'_cost'])
            else:
                x[x_name]['cost'] = 0.0
            x[x_name]['quantity'] = pulp.LpVariable(x_name, 0, None, 'Continuous')
            max_min[b_index] = {}
            if req['max_'+b_index]:
                max_min[b_index]['max'] = float(req['max_'+b_index])
            else:
                max_min[b_index]['max'] = float('inf')
            if req['min_'+b_index]:
                max_min[b_index]['min'] = float(req['min_'+b_index])
            else:
                max_min[b_index]['min'] = 0.0

    return x, max_min


def parse_result(status, x_info):

    i = 0
    sum_cost = 0.0
    results = {}
    for x_name, x_value in x_info.items():
        results[x_name] = x_value['quantity'].value()
        sum_cost += x_value['quantity'].value() * x_value['cost']
        i += 1
    if status['status'] != -1:
        results['sum_cost'] = sum_cost
        return results
    else:
        results['sum_cost'] = -1
        return results


def get_data_from_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    linear_problem_sheet = wb.get_sheet_by_name('linear problem')
    names = linear_problem_sheet['A'][1:]
    a_names = linear_problem_sheet[1][2:]
    x = {}
    i = 1
    for name in names:
        x[name.value] = {}
        j = 1
        for a_name in a_names:
            x[name.value][a_name.value] = linear_problem_sheet.cell(1 + i, 2 + j).value
            j += 1
        x[name.value]['cost'] = linear_problem_sheet.cell(1 + i, 2).value
        i += 1
    restrictive_condition_sheet = wb.get_sheet_by_name('Restrictive conditions')
    b_names = restrictive_condition_sheet['A'][1:]
    max = restrictive_condition_sheet['C'][1:]
    min = restrictive_condition_sheet['B'][1:]
    max_min = {}
    j = 0
    for b_name in b_names:
        max_min[b_name.value] = {}
        max_min[b_name.value]['max'] = max[j].value
        max_min[b_name.value]['min'] = min[j].value
        j += 1
    wb.close()

    data = {}
    x_names = x.keys()
    i = 0
    for x_name in x_names:
        data['x_name_' + chr(ord('a')+i)] = x_name
        a_names = x[x_name].keys()
        data[chr(ord('a') + i) + '_cost'] = x[x_name]['cost']
        j = 1
        for a_name in a_names:
            if a_name != 'cost' and a_name != 'quantity':
                data[chr(ord('a')+i) + str(j)] = x[x_name][a_name]
                data['max_' + str(j)] = max_min[b_names[j-1].value]['max']
                data['min_' + str(j)] = max_min[b_names[j-1].value]['min']
                data[str(j) + '_name'] = a_name
                j += 1
        i += 1
    a_names_list = list(a_names)
    a_names_list.remove('cost')
    x_names_list = list(x_names)
    data['a_names'] = dict(zip(a_names_list, a_names_list))
    data['x_names'] = x_names_list
    return data
